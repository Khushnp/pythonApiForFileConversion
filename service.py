from Models.models import Customer, CustomerBankaccount, Proxy, CustomerBulkUploadBody, BulkOperationHeader
import json
from openpyxl import load_workbook
from datetime import datetime, date

class DateEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, date):
            return obj.isoformat()
        return super().default(obj)

def process_file(file_stream, filename):
    # Load the Excel file
    wb = load_workbook(filename=file_stream, read_only=True)
    sheet = wb.active

    # Get the headers
    headers = [cell.value for cell in sheet[1]]

    # Create the JSON data
    json_data = []

    # Create the BulkOperationHeader object
    header = BulkOperationHeader(
        bankCode="80755",
        groupCode="80755",
        recordNumber=sheet.max_row - 1,  # Excluding the header row
        operationType="E",
        fileNumber="00001",
        date=datetime.now().strftime('%Y%m%d')
    )
    # Appending header first
    json_data.append(header.dict())

    # Set the values from the Excel file
    for row in sheet.iter_rows(min_row=2):
        customer = Customer(
            name=row[headers.index("Name")].value,
            surname=row[headers.index("Surname")].value,
            bankUserId=row[headers.index("BankUserId")].value,
            countyCode=row[headers.index("CountyCode")].value,
            birthDate=row[headers.index("BirthDate")].value,
            cityOfBirth=row[headers.index("CityOfBirth")].value,
            EID=row[headers.index("EID")].value,
            NID=row[headers.index("NID")].value,
            passport=row[headers.index("Passport")].value,
            economicActivityCode=row[headers.index("EconomicActivityCode")].value,
            documentIssuerCountryCode=row[headers.index("DocumentIssuerCountryCode")].value,
            documentType=row[headers.index("DocumentType")].value
        )

        bank_accounts = [
            CustomerBankaccount(
                IBAN=row[headers.index("IBAN")].value,
                currency=row[headers.index("Currency")].value,
                accountType=row[headers.index("AccountType")].value
            )
        ]

        proxies = [
            Proxy(
                type=row[headers.index("ProxyType")].value,
                value=row[headers.index("ProxyValue")].value
            )
        ]

        data = {
            "operationType": "E",
            "customer": customer.dict(),
            "mobile": row[headers.index("Mobile")].value,
            "bankAccount": [ba.dict() for ba in bank_accounts],
            "proxies": [proxy.dict() for proxy in proxies]
        }

        json_data.append(data)

    # Serialize JSON data to string using custom encoder
    json_string = "\n".join(json.dumps(data, cls=DateEncoder) for data in json_data)

    # Save the JSON data to a file
    generated_file_name = f"JIFFY.{header.groupCode}.BLKOPE.BlkOp_{header.groupCode}_{header.bankCode}_{header.date}_{header.fileNumber}.txt"
    print(f"Processing file: {filename}")
    print(f"Processing file: {generated_file_name}")
    file_path = f"F:/AutoCreated/{generated_file_name}"
    with open(file_path, "w") as json_file:
        json_file.write(json_string)

    return True
